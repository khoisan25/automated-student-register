import cognitive_face as CF
from global_variables import personGroupId
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
try: 
    from openpyxl.cell import get_column_letter, Cell, column_index_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl import cell



#from openpyxl.cell import get_column_letter, Cell, column_index_from_string
import time


#get current date
print("registering current date in report...")
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports.xlsx")
sheet = wb.get_sheet_by_name('Cse15')

def getDateColumn():
	for i in range(1, sum(1 for x in (sheet.columns)) + 1):
		col = get_column_letter(i)
		if sheet[('%s%s'% (col,'1'))].value == currentDate:
			return col
			
print("registering API on server...")		
Key = str(open('resources/APIkey.txt').read().strip())
CF.Key.set(Key)

BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0/'  
CF.BaseUrl.set(BASE_URL)

print("Connecting to database...")
connect = connect = sqlite3.connect("Face-DataBase")
c = connect.cursor()

attend = [0 for i in range(100)]	

print("ensuring file directories are present...")
currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'RegisterNow')
for filename in os.listdir(directory):
	if filename.endswith(".jpg"):
		print(filename)
		#imgurl = urllib.pathname2url(os.path.join(directory, filename))
		imgpath = os.path.join(directory, filename)
		imgurl = imgpath
		print("Starting face detection on image...")
		print(imgpath)

		res = CF.face.detect(imgurl)
		if len(res) != 1:
			print ("No face detected.")
			continue
			
		faceIds = []
		for face in res:
			faceIds.append(face['faceId'])
		res = CF.face.identify(faceIds, personGroupId)
		print (filename)
		print (res)
		for face  in res:
			if not face['candidates']:
				print ("Unknown")
			else:
				personId = face['candidates'][0]['personId']
				c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
				print ('person ID is: '+ personId)
				row = c.fetchone()

				print (row)

				if row:
					attend[int(row[0])] += 1
					print (row[1] + " recognized")
		time.sleep(2)


print("registering present students...")

for row in range(2, sum(1 for x in (sheet.columns)) + 1):
	rn = sheet['A%s'% row].value
	print(rn)
	if rn is not None:
		rn = rn[-2:]
		print(rn)
		if attend[int(rn)] != 0:
			col = getDateColumn()
			sheet['%s%s' % (col, str(row))] = 1
		
wb.save(filename = "reports.xlsx")
print("report saved.")





		
# for row in range[2, len(sheet.columns[0]) + 1]:
# 	rn = sheet.cell('A%s'% row).value
# 	if rn is not None:
# 		rn = rn[-2:]
# 		if attend[int(rn)] != 0:
# 			col = getDateColumn()
# 			sheet['%s%s' % (col, str(row))] = 1

# wb.save(filename = "reports.xlsx")













#old junk, might need to implemtnt it in the whie loop, 
#currentDir = os.path.dirname(os.path.abspath(__file__))
#imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(imgurl)
#faceIds = []
#for face in res:
 #   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
#print res
