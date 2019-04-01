from openpyxl import Workbook, load_workbook
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
import time
import os
import sqlite3

#database connection
conn = sqlite3.connect('Face-DataBase')
c = conn.cursor()
print("Connecting to database...")

#get current date
currentDate = time.strftime("%d_%m_%y")

#create a workbook and add a worksheet
if (os.path.exists('./reports.xlsx')):
    wb = load_workbook(filename = "reports.xlsx")
    sheet = wb.get_sheet_by_name('Cse15')
    # sheet[ord() + '1']
    for col_index in range(1, 100):
    	col = get_column_letter(col_index)
    	if sheet[('%s%s' % (col,1))].value is None:
    		col2 = get_column_letter(col_index - 1)
    		# print sheet.cell('%s%s'% (col2, 1)).value
    		if sheet[('%s%s' % (col2,1))].value != currentDate:
    			sheet['%s%s' % (col,1)] = currentDate
    		break
    print("Creating excel sheet...")

    #saving the file
    wb.save(filename = "reports.xlsx")
    print("Report saved.")
    	
else:
    wb = Workbook()
    dest_filename = 'reports.xlsx'
    c.execute("SELECT * FROM Students ORDER BY ID ASC")
    
    #creating worksheet and giving names to column
    ws1 = wb.active
    ws1.title = "Cse15"
    ws1.append(('Student Number', 'Name', currentDate))
    ws1.append(('', '', ''))
    print("creating worksheet...")
    print("assigning names to columns...")

    #entering students information from database
    while True:
        a = c.fetchone()
        if a == None:
            break
        else:
            ws1.append((a[2], a[1]))

    #saving the file
    wb.save(filename = dest_filename)
    print("Report saved.")
    
